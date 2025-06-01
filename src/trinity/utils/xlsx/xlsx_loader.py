from io import BytesIO

import polars as pl
from openpyxl import load_workbook


# TODO: Написать тесты.
def load_st(wb: str | BytesIO, ws_name: str, st_name: str) -> pl.DataFrame:
    try:
        wb = load_workbook(wb, data_only=True, read_only=False)
        ws = wb[ws_name]
    except KeyError as e:
        raise KeyError('Рабочий лист не найден.', ws_name) from e

    try:
        st = ws.tables[st_name]
    except KeyError as e:
        raise KeyError('Смарт-таблица не найдена.', st_name) from e

    # Определяем диапазон ячеек смарт-таблицы.
    st_range = st.ref
    st_cells = ws[st_range]

    # Извлекаем данные из ячеек смарт-таблицы.
    st_data = [[cell.value for cell in row] for row in st_cells]

    # Создаем DataFrame из данных смарт-таблицы.
    header: list[str] = st_data[0]  # Заголовок DataFrame.
    data: list[list[object]] = st_data[1:]  # Данные DataFrame.

    rows: list[dict[str, str]] = []

    # Во избежание ошибок, связанных с типизацией, преобразуем значения в строки.
    try:
        for _, row in enumerate(data, start=1):
            row = {
                c_name: (str(c_value) if c_value is not None else '')
                for c_name, c_value in zip(header, row, strict=True)
            }
            rows.append(row)
    except ValueError as e:
        raise ValueError('Ошибка при сборке данных DataFrame.', e) from e

    df = pl.DataFrame(rows)

    return df
