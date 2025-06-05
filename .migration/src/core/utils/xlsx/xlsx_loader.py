import logging
from io import BytesIO

import polars as pl
from openpyxl import load_workbook

log = logging.getLogger(__name__)


def load_smart_table(workbook: str | BytesIO, worksheet: str, smart_table: str) -> pl.DataFrame:
    """
    Загружает смарт-таблицу из файла Excel и возвращает ее как polars DataFrame.

    Args:
        workbook: Путь к файлу Excel или объект BytesIO.
        worksheet: Имя листа, содержащего смарт-таблицу.
        smart_table: Имя смарт-таблицы.

    Returns:
        Данные смарт-таблицы в виде DataFrame.
    """
    try:
        wb = load_workbook(workbook, data_only=True, read_only=False)
        ws = wb[worksheet]
    except KeyError:
        log.exception('Не удалось загрузить смарт-таблицу. Лист [%s] не найден.', worksheet)
        raise

    try:
        st = ws.tables[smart_table]
    except KeyError:
        log.exception('Не удалось загрузить смарт-таблицу. Смарт-таблица [%s] не найдена.', smart_table)
        raise

    # Определяем диапазон ячеек смарт-таблицы
    st_range = st.ref
    st_cells = ws[st_range]

    # Сбор данных из ячеек смарт-таблицы
    st_data = [[cell.value for cell in row] for row in st_cells]  # type: ignore

    # Преобразование данных в polars DataFrame
    header: list[str] = st_data[0]
    data: list[list[object]] = st_data[1:]

    # Создаем список словарей для последующего преобразования в DataFrame
    rows: list[dict[str, str]] = []

    # Во избежание ошибок, связанных с типизацией, преобразуем значения в строки
    for _, row in enumerate(data, start=1):
        row = {c_name: (str(c_value) if c_value is not None else '') for c_name, c_value in zip(header, row)}
        rows.append(row)

    df = pl.DataFrame(rows)

    return df
